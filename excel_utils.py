import openpyxl
import numpy as np
from openpyxl.styles import Alignment
import os

# Función para calcular estadísticas a partir de un archivo de consultas numéricas
def calcular_estadisticas():
    # Ruta del archivo de consultas numéricas
    nombre_archivo = os.path.join('Consultas', 'consultas_numericas.txt')

    try:
        # Abrir el archivo en modo lectura
        with open(nombre_archivo, "r") as archivo:
            # Leer todas las líneas del archivo
            lineas = archivo.readlines()

        # Diccionario para almacenar datos estadísticos por parámetro
        datos_por_parametro = {}

        # Variables para rastrear el parámetro actual durante el procesamiento del archivo
        parametro_actual = None  

        # Iterar sobre las líneas del archivo
        for linea in lineas:
            # Dividir la línea en dos partes utilizando el primer ":" como separador
            partes = [parte.strip() for parte in linea.split(":", 1)]
            
            if len(partes) == 2:
                # Asignar la primera parte como el parámetro actual y la segunda como el resto de la línea
                parametro_actual, resto = partes
                resto = resto.strip().split()
                
                # Verificar si hay al menos un elemento en el resto
                if len(resto) > 0:
                    try:
                        # Intentar convertir el primer elemento del resto a un valor numérico (float)
                        valor = float(resto[0])  
                    except ValueError:
                        # Ignorar la línea si no se puede convertir a un valor numérico
                        print(f"Ignorando línea: {linea}")
                        continue  

                    # Agregar el valor numérico al diccionario por parámetro
                    if parametro_actual not in datos_por_parametro:
                        datos_por_parametro[parametro_actual] = []

                    datos_por_parametro[parametro_actual].append(valor)
            elif parametro_actual and not linea.strip():  
                # Restablecer el parámetro actual a None cuando se encuentre una línea en blanco
                parametro_actual = None
            elif parametro_actual:
                # Ignorar líneas adicionales si hay un parámetro actual
                print(f"Ignorando línea: {linea}")

        # Calcular estadísticas para cada parámetro
        for parametro, datos in datos_por_parametro.items():
            if datos:
                promedio = np.mean(datos)
                maximo = np.max(datos)
                minimo = np.min(datos)

                print(f"\nEstadísticas para {parametro}:")
                print(f"Promedio: {promedio}")
                print(f"Máximo: {maximo}")
                print(f"Mínimo: {minimo}")
            else:
                print(f"\nNo hay datos numéricos para {parametro} en el archivo.")

    except FileNotFoundError:
        # Manejar el caso en que el archivo no existe o está vacío
        print(f"El archivo {nombre_archivo} no existe o está vacío.")
    except Exception as e:
        # Manejar otras excepciones durante la lectura del archivo
        print("Error al leer los datos:", str(e))
    
    # Devolver el diccionario con datos estadísticos
    return datos_por_parametro

# Función para crear un nuevo libro de Excel
def crear_libro():
    libro = openpyxl.Workbook()
    return libro

# Función para guardar estadísticas en una hoja de un libro de Excel
def guardar_estadisticas(libro, datos_por_parametro):
    # Obtener la hoja activa del libro
    hoja = libro.active
    hoja.title = "Estadisticas"  # Establecer el título de la hoja

    fila_actual = 1

    # Iterar sobre los parámetros y sus datos en el diccionario
    for parametro, datos in datos_por_parametro.items():
        hoja.cell(row=fila_actual, column=1, value=f"Estadísticas para {parametro}")
        fila_actual += 1
        hoja.cell(row=fila_actual, column=1, value="Promedio:")
        hoja.cell(row=fila_actual, column=2, value=np.mean(datos))
        fila_actual += 1
        hoja.cell(row=fila_actual, column=1, value="Máximo:")
        hoja.cell(row=fila_actual, column=2, value=np.max(datos))
        fila_actual += 1
        hoja.cell(row=fila_actual, column=1, value="Mínimo:")
        hoja.cell(row=fila_actual, column=2, value=np.min(datos))
        fila_actual += 2  

    # Alinear las celdas en la hoja
    for fila in hoja.iter_rows(min_row=1, max_col=2, max_row=hoja.max_row):
        for celda in fila:
            celda.alignment = Alignment(horizontal='center', vertical='center')

# Función para guardar un libro de Excel en una carpeta específica
def guardar_libro(libro, nombre_archivo="estadisticas.xlsx"):
    # Ruta completa del archivo de salida, incluyendo la carpeta "Reporte"
    nombre_archivo = os.path.join('Reporte', nombre_archivo)
    
    # Guardar el libro de Excel en el archivo especificado
    libro.save(nombre_archivo)
    
    # Imprimir un mensaje indicando que el libro se ha guardado
    print(f"Libro de Excel guardado como {nombre_archivo}")
